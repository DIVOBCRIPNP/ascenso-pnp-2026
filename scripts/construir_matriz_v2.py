"""
Construye data/patrones.json v3 fusionando:
  1. El banco `data/preguntas.json` (1500 preguntas, base de identidad).
  2. La distribución metodológica del usuario (P0/G1..G6, MD).
  3. La Matriz de coincidencias del usuario (XLSX):
       - 01_Matriz_1500: familia, subfamilia, palabras cambiantes, tipos, dificultad,
                         técnica, observación de memorización, dificultad, ubicación.
       - 02_Ranking_familias: cantidades por familia.
       - 04_Respuestas_repetidas: grupos de preguntas con respuesta idéntica.
       - 05_Preguntas_espejo: familias por inicio compartido.
       - 06_Ruta_estudio: niveles 0-10 con método y prioridad.

Salida: data/patrones.json (v3)
  {
    version: 3,
    meta: {total, familias_conteo, tipos_coincidencia_conteo, familias_mismo_inicio},
    grupos: [ {id:P0..G6, label, ...} ],           // ruta metodológica
    familias: [ {id, label, cantidad, como_estudiar, ns:[...]}, ],
    ruta: [ {nivel, grupo, contiene, metodo, prioridad} ],
    respuestas_repetidas: [ {respuesta, cantidad, ns:[...], familia, tecnica} ],
    inicios: {clave: [ns...]},
    por_pregunta: {
      n: {
        grupo, familia, subfamilia,
        palabras_cambian: [str],       // array de tokens a resaltar
        tipos: [str],                  // etiquetas de coincidencia
        dificultad, tecnica, observacion,
        relacionadas: [ns...],         // por respuesta idéntica
      }
    }
  }
"""
import json, os, re
from collections import defaultdict, Counter
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BANK = os.path.join(ROOT, "data", "preguntas.json")
OUT  = os.path.join(ROOT, "data", "patrones.json")

XLSX_MATRIZ = "/Users/javo/Downloads/matriz_coincidencias_pregunta_respuesta_001_1500.xlsx"
XLSX_META   = "/Users/javo/Downloads/patrones_banco_preguntas_lote_001_300.xlsx"  # base P0..G6 (grupos)
MD_RUTA     = "/Users/javo/Downloads/distribucion_metodologica_estudio_1500_preguntas.md"

with open(BANK) as f: DATA = json.load(f)
DATA.sort(key=lambda q: q.get("n", 0))
print(f"Banco: {len(DATA)}")

# ------- Matriz de coincidencias (XLSX principal) -------
wb = openpyxl.load_workbook(XLSX_MATRIZ, data_only=True)

# 01_Matriz_1500 → por_pregunta
ws = wb["01_Matriz_1500"]
h  = [c.value for c in ws[1]]; ix = {k:i for i,k in enumerate(h)}
por_pregunta = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    n = row[ix["N°"]]
    if not n: continue
    try: n = int(n)                       # normaliza a int: el XLSX lo trae como string
    except: continue
    def val(k, default=""):
        return (row[ix[k]] or default) if k in ix else default
    palabras = str(val("Palabra clave que cambia") or "")
    # separo por coma o "/", tokenizo palabras significativas
    palabras_arr = [t.strip() for t in re.split(r"[,/;]", palabras) if t.strip() and len(t.strip())>=3]
    tipos = str(val("Tipo de coincidencia") or "")
    tipos_arr = [t.strip() for t in tipos.split(";") if t.strip()]
    por_pregunta[n] = {
        "familia": str(val("Familia temática principal") or "").strip(),
        "subfamilia": str(val("Subfamilia") or "").strip(),
        "palabras_cambian": palabras_arr,
        "tipos": tipos_arr,
        "dificultad": str(val("Nivel de dificultad") or "").strip(),
        "tecnica": str(val("Técnica de estudio recomendada") or "").strip(),
        "observacion": str(val("Observación para memorizar") or "").strip(),
        "estado_rd": str(val("Estado RD") or "").strip(),
        "patron_pregunta": str(val("Patrón en la pregunta") or "").strip(),
        "patron_respuesta": str(val("Patrón en la respuesta") or "").strip(),
        "relacionadas": [],
    }
print(f"  01_Matriz_1500: {len(por_pregunta)} filas ingeridas")

# 02_Ranking_familias
familias = []
ws = wb["02_Ranking_familias"]
h  = [c.value for c in ws[1]]; ix = {k:i for i,k in enumerate(h)}
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[ix["Familia temática"]]: continue
    familias.append({
        "id": row[ix["Familia temática"]],
        "label": row[ix["Familia temática"]],
        "cantidad": int(row[ix["Cantidad"]] or 0),
        "como_estudiar": row[ix["Cómo estudiarlo"]] or "",
        "observacion": row[ix["Observación"]] or "",
        "ns": [],  # se llena abajo
    })
# rellena ns por familia
familia_by_label = {f["id"]: f for f in familias}
for n, info in por_pregunta.items():
    fam = info["familia"]
    if fam in familia_by_label:
        familia_by_label[fam]["ns"].append(n)
for f in familias:
    f["ns"].sort()
print(f"  02_Ranking_familias: {len(familias)} familias")

# 04_Respuestas_repetidas
respuestas_repetidas = []
ws = wb["04_Respuestas_repetidas"]
h  = [c.value for c in ws[1]]; ix = {k:i for i,k in enumerate(h)}
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[ix["Respuesta oficial repetida"]]: continue
    asoc = str(row[ix["Preguntas asociadas"]] or "")
    ns = [int(x) for x in re.findall(r"\b\d+\b", asoc) if 1 <= int(x) <= 1500]
    respuestas_repetidas.append({
        "respuesta": row[ix["Respuesta oficial repetida"]],
        "cantidad": int(row[ix["Cantidad"]] or 0),
        "ns": ns,
        "familia": row[ix["Familia temática"]] or "",
        "tecnica": row[ix["Técnica de estudio"]] or "",
    })
# marca en por_pregunta.relacionadas los hermanos por respuesta idéntica
for grupo in respuestas_repetidas:
    ns_grupo = set(grupo["ns"])
    for n in grupo["ns"]:
        if n in por_pregunta:
            hermanos = sorted(ns_grupo - {n})
            por_pregunta[n]["relacionadas"] = hermanos
print(f"  04_Respuestas_repetidas: {len(respuestas_repetidas)} grupos")

# 05_Preguntas_espejo (mismo inicio) — construimos también nuestro índice basado en 7 primeras palabras
def norm(s):
    s = (s or "").upper()
    s = re.sub(r"[^A-ZÁÉÍÓÚÑ0-9\s]", " ", s)
    s = re.sub(r"[ÁÉÍÓÚ]", lambda m: "AEIOU"["ÁÉÍÓÚ".index(m.group(0))], s)
    return re.sub(r"\s+", " ", s).strip()

def inicio(s, n=7):
    return " ".join(norm(s).split()[:n])

buckets = defaultdict(list)
for q in DATA:
    k = inicio(q.get("pregunta",""))
    if k: buckets[k].append(q["n"])
inicios = {k: sorted(v) for k, v in buckets.items() if len(v) >= 2}
print(f"  Familias 'mismo inicio': {len(inicios)}")

# 06_Ruta_estudio
ruta = []
ws = wb["06_Ruta_estudio"]
h  = [c.value for c in ws[1]]; ix = {k:i for i,k in enumerate(h)}
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[ix["Nivel"]] is None: continue
    ruta.append({
        "nivel": row[ix["Nivel"]],
        "grupo": row[ix["Grupo de estudio"]] or "",
        "contiene": row[ix["Qué contiene"]] or "",
        "metodo": row[ix["Método"]] or "",
        "prioridad": row[ix["Prioridad"]] or "",
    })
print(f"  06_Ruta_estudio: {len(ruta)} niveles")

# ------- Ruta metodológica P0..G6 (del MD del usuario) -------
def parse_grupos_md(path):
    txt = open(path).read()
    grupos = []
    for m in re.finditer(r"### (P0|G\d)([^\n]*)\((\d+)\s*preguntas?\)\s*\n((?:[^\n]|\n(?!###))*)", txt):
        gid = m.group(1); label = m.group(2).strip(); expected = int(m.group(3))
        nums = sorted(set(int(x) for x in re.findall(r"\b\d+\b", m.group(4)) if 1 <= int(x) <= 1500))
        grupos.append({"id": gid, "label": label, "esperado": expected, "n": nums})
    meta_tbl = {
        "P0": ("Especial",     "Memorizar la versión final oficial antes de mezclar con el banco base."),
        "G1": ("Fácil",         "Tarjetas concepto/definición → respuesta."),
        "G2": ("Fácil-media",   "Tabla acción → autoridad competente."),
        "G3": ("Media",         "Tabla número → tema → respuesta."),
        "G4": ("Media",         "Agrupar por inicio y estudiar el complemento que cambia."),
        "G5": ("Alta",          "Comparar opciones; subrayar la palabra que cambia."),
        "G6": ("Muy alta",      "Leer primero no/salvo/excepto/incorrecta antes de responder."),
    }
    for g in grupos:
        d, c = meta_tbl.get(g["id"], ("", ""))
        g["dificultad"] = d
        g["como_estudiar"] = c
    return grupos

grupos = parse_grupos_md(MD_RUTA)
grupo_by_n = {}
for g in grupos:
    for n in g["n"]:
        if n in grupo_by_n and grupo_by_n[n] != g["id"]:
            grupo_by_n[n] += "+" + g["id"]
        else:
            grupo_by_n[n] = g["id"]
# adjunta grupo en cada pregunta
for q in DATA:
    n = q["n"]
    if n not in por_pregunta:
        por_pregunta[n] = {"familia":"", "subfamilia":"", "palabras_cambian":[], "tipos":[],
                           "dificultad":"", "tecnica":"", "observacion":"", "relacionadas":[]}
    por_pregunta[n]["grupo"] = grupo_by_n.get(n, "")

# ------- meta agregada -------
tipos_conteo = Counter()
familia_conteo = Counter()
for info in por_pregunta.values():
    for t in info.get("tipos", []): tipos_conteo[t] += 1
    if info.get("familia"): familia_conteo[info["familia"]] += 1

out = {
    "version": 3,
    "meta": {
        "total": len(DATA),
        "familias_conteo": dict(familia_conteo.most_common()),
        "tipos_conteo": dict(tipos_conteo.most_common()),
        "familias_mismo_inicio": len(inicios),
    },
    "grupos": grupos,
    "familias": sorted(familias, key=lambda f: -f["cantidad"]),
    "ruta": ruta,
    "respuestas_repetidas": respuestas_repetidas,
    "inicios": inicios,
    "por_pregunta": por_pregunta,
}

with open(OUT, "w") as f:
    json.dump(out, f, ensure_ascii=False)
print(f"\n✓ Escrito {OUT}  ({os.path.getsize(OUT)//1024} KB)")
print(f"  familias con cantidades: {list(familia_conteo.most_common(4))}")
print(f"  tipos con conteos: {list(tipos_conteo.most_common(4))}")
