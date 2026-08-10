"""
Construye data/patrones.json fusionando:
  1. Tu análisis manual del lote 001-300 (XLSX 1) — patrones por pregunta.
  2. Tu análisis manual de los lotes 301-1500 (XLSX 2) — patrones por pregunta.
  3. Tu distribución metodológica en 6 grupos + P0 (MD) — grupo pedagógico por pregunta.

Salida:
  data/patrones.json
    - meta: totales por grupo y por patrón
    - grupos: definiciones (id, label, dificultad, cómo estudiarlo, listado de n°)
    - por_pregunta: { n: { grupo, tags, similitud, palabras_cambian, palabras_clave, regla, rd } }
    - inicios: familias de "mismo inicio" con las preguntas relacionadas
"""
import json, os, re, sys
from collections import Counter, defaultdict

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BANK = os.path.join(ROOT, "data", "preguntas.json")
OUT  = os.path.join(ROOT, "data", "patrones.json")

XLSX1 = "/Users/javo/Downloads/patrones_banco_preguntas_lote_001_300.xlsx"
XLSX2 = "/Users/javo/Downloads/patrones_banco_preguntas_lotes_301_1500.xlsx"
MD_RUTA = "/Users/javo/Downloads/distribucion_metodologica_estudio_1500_preguntas.md"

import openpyxl

# ------- 1. cargar el banco -------
with open(BANK) as f: DATA = json.load(f)
DATA.sort(key=lambda q: q.get("n", 0))
byn = {q["n"]: q for q in DATA}

# ------- 2. cargar patrones por pregunta (dos XLSX del usuario) -------
por_pregunta = {}

def add_p1(n, row, ix):
    """XLSX 1 (001-300) - esquema del primer archivo."""
    tags = []
    if row[ix["Respuesta más larga"]] == "Sí": tags.append("respuesta_larga_estricta")
    if row[ix["Respuesta más larga"]] == "Empate": tags.append("respuesta_larga_empate")
    if row[ix["Opciones casi iguales"]] == "Sí": tags.append("opciones_casi_iguales")
    if row[ix["Negación/excepción"]] == "Sí": tags.append("negacion_o_excepcion")
    pdet = (row[ix["Patrón detectado"]] or "").lower()
    if "marcar la incorrecta" in pdet: tags.append("marcar_la_incorrecta")
    por_pregunta[n] = {
        "tags": sorted(set(tags)),
        "similitud": row[ix["Similitud máx."]] or 0.0,
        "par_max": row[ix["Par de opciones"]] or "",
        "regla": row[ix["Regla de estudio"]] or "",
        "palabras_cambian": "",
        "palabras_clave": "",
    }

def add_p2(n, row, ix):
    """XLSX 2 (301-1500) - esquema del segundo archivo (más rico)."""
    tags = []
    v = row[ix["¿Respuesta más larga?"]]
    if v == "Sí, clara": tags.append("respuesta_larga_estricta")
    if v == "Sí, empatada": tags.append("respuesta_larga_empate")
    if row[ix["¿Opciones similares?"]] == "Sí": tags.append("opciones_casi_iguales")
    if row[ix["¿Negativa/excepción?"]] == "Sí": tags.append("negacion_o_excepcion")
    if row[ix["¿Inicio repetido?"]] == "Sí": tags.append("mismo_inicio")
    pdet = (row[ix["Patrón detectado"]] or "").lower()
    if "marcar la incorrecta" in pdet: tags.append("marcar_la_incorrecta")
    por_pregunta[n] = {
        "tags": sorted(set(tags)),
        "similitud": row[ix["Similitud máx."]] or 0.0,
        "palabras_cambian": row[ix["Palabras que cambian"]] or "",
        "palabras_clave": row[ix["Palabras clave"]] or "",
        "regla": row[ix["Regla de estudio"]] or "",
        "estado_rd": row[ix["Estado RD"]] or "",
    }

wb1 = openpyxl.load_workbook(XLSX1, data_only=True)
ws1 = wb1["Preguntas 001-300"]
h1 = [c.value for c in ws1[1]]; ix1 = {k:i for i,k in enumerate(h1)}
for row in ws1.iter_rows(min_row=2, values_only=True):
    n = row[ix1["N°"]]
    if n: add_p1(n, row, ix1)
print(f"XLSX1 (001-300): {sum(1 for n in por_pregunta if n<=300)} preguntas")

wb2 = openpyxl.load_workbook(XLSX2, data_only=True)
for sname in ["301-600","601-900","901-1200","1201-1500"]:
    ws = wb2[sname]
    h = [c.value for c in ws[1]]; ix = {k:i for i,k in enumerate(h)}
    for row in ws.iter_rows(min_row=2, values_only=True):
        n = row[ix["N°"]]
        if n: add_p2(n, row, ix)
print(f"Total por_pregunta: {len(por_pregunta)}")

# También agrego mismo_inicio para 001-300 (el XLSX1 no lo trae por fila)
def norm(s):
    s=(s or "").upper()
    s=re.sub(r"[^A-ZÁÉÍÓÚÑ0-9\s]"," ",s)
    s=re.sub(r"[ÁÉÍÓÚ]", lambda m:"AEIOU"["ÁÉÍÓÚ".index(m.group(0))], s)
    return re.sub(r"\s+"," ",s).strip()

def inicio(s, n=7): return " ".join(norm(s).split()[:n])

buckets = defaultdict(list)
for q in DATA:
    k = inicio(q.get("pregunta",""))
    if k: buckets[k].append(q["n"])
inicios = {k: sorted(v) for k,v in buckets.items() if len(v)>=2}
mismo_inicio_set = {n for arr in inicios.values() for n in arr}
for n in mismo_inicio_set:
    if n in por_pregunta and "mismo_inicio" not in por_pregunta[n]["tags"]:
        por_pregunta[n]["tags"] = sorted(set(por_pregunta[n]["tags"]) | {"mismo_inicio"})
print(f"Familias 'mismo inicio': {len(inicios)}")

# ------- 3. parsear la distribución metodológica en 6 grupos + P0 -------
def parse_grupos(md_path):
    txt = open(md_path).read()
    grupos = []
    for m in re.finditer(r"### (P0|G\d)([^\n]*)\((\d+)\s*preguntas?\)\s*\n((?:[^\n]|\n(?!###))*)", txt):
        gid = m.group(1); label = m.group(2).strip(); expected = int(m.group(3))
        body = m.group(4)
        # une líneas y separa números
        nums = re.findall(r"\b\d+\b", body)
        nums = [int(x) for x in nums]
        # protección contra números "de renglón" — todos los válidos son 1..1500
        nums = sorted(set(x for x in nums if 1 <= x <= 1500))
        grupos.append({"id": gid, "label": label, "esperado": expected, "n": nums})
    # dificultad y "cómo estudiarlo" salen de la tabla superior
    meta_tbl = {
        "P0": ("Especial",     "Memorizar la versión final oficial antes de mezclar con el banco base."),
        "G1": ("Fácil",         "Tarjetas concepto/definición → respuesta."),
        "G2": ("Fácil-media",   "Tabla acción → autoridad competente."),
        "G3": ("Media",         "Tabla número → tema → respuesta."),
        "G4": ("Media",         "Agrupar por inicio y estudiar el complemento que cambia."),
        "G5": ("Alta",          "Comparar opciones; subrayar la palabra que cambia."),
        "G6": ("Muy alta",      "Leer primero no/salvo/excepto/incorrecta antes de responder."),
    }
    for g in grupos:
        d, c = meta_tbl.get(g["id"], ("",""))
        g["dificultad"] = d
        g["como_estudiar"] = c
    return grupos

grupos = parse_grupos(MD_RUTA)
for g in grupos:
    marca = "✓" if len(g["n"]) == g["esperado"] else f"⚠ ({len(g['n'])} vs esperado {g['esperado']})"
    print(f"  {g['id']:>3} · {g['label'][:35]:<35} {len(g['n']):>4} {marca}")

# marca en cada pregunta a qué grupo pertenece
grupo_by_n = {}
for g in grupos:
    for n in g["n"]:
        # P0 puede coexistir con otro grupo; guardamos ambos
        if n in grupo_by_n and grupo_by_n[n] != g["id"]:
            grupo_by_n[n] = grupo_by_n[n] + "+" + g["id"]
        else:
            grupo_by_n[n] = g["id"]

# ------- 4. arma salida -------
# asegura entrada para todas las 1500 (aunque no tengan patrones)
for q in DATA:
    n = q["n"]
    if n not in por_pregunta:
        por_pregunta[n] = {"tags": [], "similitud": 0.0, "regla": "", "palabras_cambian":"", "palabras_clave":""}
    por_pregunta[n]["grupo"] = grupo_by_n.get(n, "")

# meta: conteos por patrón
tag_counts = Counter()
for info in por_pregunta.values():
    for t in info["tags"]: tag_counts[t] += 1

out = {
    "version": 2,
    "meta": {
        "total": len(DATA),
        "por_patron": dict(tag_counts),
        "familias_mismo_inicio": len(inicios),
    },
    "grupos": grupos,
    "inicios": inicios,
    "por_pregunta": por_pregunta,
}

with open(OUT, "w") as f:
    json.dump(out, f, ensure_ascii=False)
print(f"\n✓ Escrito {OUT}  ({os.path.getsize(OUT)//1024} KB)")
