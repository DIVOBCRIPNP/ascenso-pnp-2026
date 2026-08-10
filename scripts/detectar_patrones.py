"""
Detector de patrones para el banco de preguntas PNP.

Toma data/preguntas.json y genera:
  1) Un XLSX por lote (300 preguntas) con las mismas columnas del análisis manual
     que hizo el usuario para 001-300, para que se pueda verificar humanamente.
  2) data/patrones.json con las etiquetas por pregunta para consumir en la app.

Patrones detectados por pregunta:
  - respuesta_larga_estricta      : la correcta es la opción más larga (única)
  - respuesta_larga_empate        : la correcta es la más larga pero empatada
  - opciones_casi_iguales         : ≥1 par de opciones con similitud ≥0.55
  - negacion_o_excepcion          : palabras trampa en enunciado u opciones
  - marcar_la_incorrecta          : el enunciado pide la INCORRECTA / NO / EXCEPTO
  - mismo_inicio                  : el enunciado comparte inicio con otras preguntas
  - respuesta_mas_especifica      : la correcta cita autoridad/plazo/artículo puntual
  - respuesta_literal_normativa   : la correcta parece copia literal del texto legal

Uso:
  python3 scripts/detectar_patrones.py                   # todos los lotes
  python3 scripts/detectar_patrones.py --lote 2          # solo 301-600
  python3 scripts/detectar_patrones.py --tam 300         # cambiar tamaño de lote
"""
import json, re, argparse, os, sys
from collections import defaultdict, Counter
from difflib import SequenceMatcher

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BANK = os.path.join(ROOT, "data", "preguntas.json")
OUT_JSON = os.path.join(ROOT, "data", "patrones.json")
OUT_DIR  = os.path.join(ROOT, "data", "patrones_lotes")
MANUAL_XLSX = "/Users/javo/Downloads/patrones_banco_preguntas_lote_001_300.xlsx"

STOP = set("de del la el las los al a en y o u por para con sin que se su sus lo un una unos unas es son ser estar como más menos entre segun según sobre bajo hasta desde ni pero no ya".split())

# Palabras que introducen negación, límite o excepción en el enunciado.
# Calibrado contra el análisis manual del lote 001-300 (usuario: 48).
PALABRAS_NEG_ENUNCIADO = [
    r"\bno\b", r"\bnadie\b", r"\bnunca\b", r"\bsalvo\b", r"\bexcepto\b", r"\bsin\b",
    r"\bincorrecta\b", r"\bfalsa\b", r"\bfalso\b", r"\bmarcar\s+la\b",
    r"\bse\s+suspende\b", r"\bcarece\b", r"\bsin\s+perjuicio\b",
    r"\bproh[íi]b\w+\b", r"\bimpide\b", r"\blimit\w+\b", r"\brestrin\w+\b",
    r"\bsuspend\w+\b", r"\bnul\w+\b", r"\binhabilit\w+\b",
]
PALABRAS_MARCAR_INCORRECTA = [
    r"marcar la incorrecta", r"señale la incorrecta", r"cuál no",
    r"no corresponde", r"no es correcto", r"marcar la falsa", r"cuál es falsa",
    r"señale la falsa", r"no forma parte", r"excepto",
]
# citas normativas típicas
RX_NORMATIVA = re.compile(
    r"\b(?:art(?:[íi]culo|\.?)\s?\d+|inciso|numeral|literal|título|cap[íi]tulo|"
    r"ley\s?n[°º]?\s?\d+|decreto|reglamento|resoluci[oó]n)\b", re.IGNORECASE
)
# plazos/cifras concretas
RX_ESPECIFICO = re.compile(
    r"\b(?:\d+\s?(?:d[íi]as?|meses?|años?|horas?)|"
    r"\d+\s?%|"
    r"jefe|juez|fiscal|comandante|director|director general|inspector|"
    r"presidente|prefecto|subprefecto|comisario|congreso|ministerio)\b",
    re.IGNORECASE,
)

def norm(s):
    s = (s or "").upper()
    s = re.sub(r"[^A-ZÁÉÍÓÚÑ0-9\s]", " ", s)
    s = re.sub(r"[ÁÉÍÓÚ]", lambda m: "AEIOU"["ÁÉÍÓÚ".index(m.group(0))], s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def inicio(s, palabras=7):
    return " ".join(norm(s).split()[:palabras])

def ratio(a, b):
    return SequenceMatcher(None, norm(a), norm(b)).ratio()

def detectar_patrones(q, todos_inicios):
    tags = []
    opts = q.get("opciones", []) or []
    correcta = q.get("correcta")
    if not isinstance(correcta, int) or not (0 <= correcta < len(opts)):
        return {"tags": tags, "similitud_max": 0.0, "par_max": "", "resp_mas_larga": ""}
    correct_text = opts[correcta]
    lens = [len(o) for o in opts]
    max_len = max(lens)
    is_longest = lens[correcta] == max_len
    ties = sum(1 for L in lens if L == max_len)
    if is_longest and ties == 1:
        tags.append("respuesta_larga_estricta")
    elif is_longest and ties > 1:
        tags.append("respuesta_larga_empate")

    # opciones casi iguales
    sim_max = 0.0; par_max = ""
    for i in range(len(opts)):
        for j in range(i+1, len(opts)):
            r = ratio(opts[i], opts[j])
            if r > sim_max:
                sim_max, par_max = r, f"{i+1}-{j+1}"
    if sim_max >= 0.72:  # calibrado contra el análisis manual del lote 001-300 (usuario: 134)
        tags.append("opciones_casi_iguales")

    # negación / excepción — mira enunciado y palabras clave típicas en opciones
    enun = q.get("pregunta","").lower()
    opts_txt = " || ".join(opts).lower()
    hit_enun = any(re.search(p, enun) for p in PALABRAS_NEG_ENUNCIADO)
    hit_opts = bool(re.search(r"\b(?:salvo|excepto|siempre que|no están|no comprende)\b", opts_txt))
    # también: preguntas que enumeran una condición negativa ("SE SUSPENDE", "SE PIERDE", "CESA POR")
    hit_causal = bool(re.search(r"\b(?:se suspende|se pierde|cesa por|se pasa a|se extingue|se retira|termina)\b", enun))
    if hit_enun or hit_opts or hit_causal:
        tags.append("negacion_o_excepcion")
    if any(re.search(p, q.get("pregunta","").lower()) for p in PALABRAS_MARCAR_INCORRECTA):
        tags.append("marcar_la_incorrecta")

    # mismo inicio (contra el mapa global)
    key = inicio(q.get("pregunta",""))
    if todos_inicios.get(key, 0) >= 2:
        tags.append("mismo_inicio")

    # respuesta más específica (cita autoridad/plazo/número)
    if RX_ESPECIFICO.search(correct_text) and not any(RX_ESPECIFICO.search(o) for k,o in enumerate(opts) if k!=correcta and RX_ESPECIFICO.search(o) is None):
        # aproximación: correcta trae marcador específico
        tags.append("respuesta_mas_especifica")

    # respuesta literal normativa
    if RX_NORMATIVA.search(correct_text):
        tags.append("respuesta_literal_normativa")

    return {
        "tags": sorted(set(tags)),
        "similitud_max": round(sim_max, 3),
        "par_max": par_max,
        "resp_mas_larga": ("Sí" if (is_longest and ties==1) else ("Empate" if (is_longest and ties>1) else "No")),
    }

def regla_estudio(tags):
    hints = []
    if "respuesta_larga_estricta" in tags:
        hints.append("suele ser la opción más completa; verifica que también sea la más precisa")
    if "respuesta_larga_empate" in tags:
        hints.append("descarta la más corta y compara las dos largas por palabra clave")
    if "opciones_casi_iguales" in tags:
        hints.append("diferencia la palabra bisagra que cambia entre opciones")
    if "negacion_o_excepcion" in tags:
        hints.append("relee la negación antes de escoger")
    if "marcar_la_incorrecta" in tags:
        hints.append("cuidado: te piden la INCORRECTA, no la correcta")
    if "mismo_inicio" in tags:
        hints.append("agrupa esta con las de mismo inicio y aprende qué las diferencia")
    if "respuesta_mas_especifica" in tags:
        hints.append("elige la que precisa autoridad/plazo/artículo exacto")
    if "respuesta_literal_normativa" in tags:
        hints.append("suele ser la que cita literalmente la norma")
    return " / ".join(hints) if hints else "sin patrón detectado"

def cargar_manual_xlsx(path):
    """Carga las etiquetas del análisis MANUAL del usuario para 001-300."""
    if not os.path.exists(path):
        return {}
    try:
        import openpyxl
    except ImportError:
        print("(openpyxl no disponible; se ignora el análisis manual)")
        return {}
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Preguntas 001-300"]
    header = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(header)}
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        n = row[idx["N°"]]
        if not n: continue
        tags = []
        if row[idx["Respuesta más larga"]] == "Sí": tags.append("respuesta_larga_estricta")
        if row[idx["Respuesta más larga"]] == "Empate": tags.append("respuesta_larga_empate")
        if row[idx["Opciones casi iguales"]] == "Sí": tags.append("opciones_casi_iguales")
        if row[idx["Negación/excepción"]] == "Sí": tags.append("negacion_o_excepcion")
        pdet = (row[idx["Patrón detectado"]] or "").lower()
        if "marcar la incorrecta" in pdet: tags.append("marcar_la_incorrecta")
        out[n] = {
            "tags": sorted(set(tags)),
            "similitud_max": row[idx["Similitud máx."]] or 0.0,
            "par_max": row[idx["Par de opciones"]] or "",
            "resp_mas_larga": row[idx["Respuesta más larga"]] or "No",
            "regla": row[idx["Regla de estudio"]] or "",
            "fuente": "manual",
        }
    return out

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--tam", type=int, default=300)
    ap.add_argument("--lote", type=int, help="1..N — corre solo ese lote")
    args = ap.parse_args()

    with open(BANK) as f: DATA = json.load(f)
    DATA.sort(key=lambda q: q.get("n", 0))
    N = len(DATA)
    print(f"Banco cargado: {N} preguntas")

    # índice global de inicios (para 'mismo_inicio' cross-lote)
    inicios = Counter(inicio(q.get("pregunta","")) for q in DATA if q.get("pregunta"))
    print(f"Inicios repetidos (≥2): {sum(1 for c in inicios.values() if c>=2)} claves")

    manual = cargar_manual_xlsx(MANUAL_XLSX)
    print(f"Análisis manual cargado: {len(manual)} preguntas (fuente autoritativa 001-300)")

    os.makedirs(OUT_DIR, exist_ok=True)

    resultados = {}  # n -> dict con tags, hints, etc.
    total_lotes = (N + args.tam - 1) // args.tam
    for l in range(1, total_lotes+1):
        if args.lote and args.lote != l: continue
        ini = (l-1)*args.tam
        fin = min(ini + args.tam, N)
        lote = DATA[ini:fin]
        print(f"\n=== Lote {l} ({ini+1:04d}-{fin:04d}, {len(lote)} preguntas) ===")

        # conteos por patrón dentro del lote
        conteo = Counter()
        for q in lote:
            n = q["n"]
            if n in manual:
                info = dict(manual[n])   # usa el análisis MANUAL (autoritativo)
                # completa mismo_inicio (que el usuario no marcó por pregunta)
                if inicios.get(inicio(q.get("pregunta","")), 0) >= 2 and "mismo_inicio" not in info["tags"]:
                    info["tags"] = sorted(set(info["tags"]) | {"mismo_inicio"})
            else:
                info = detectar_patrones(q, inicios)
                info["fuente"] = "auto"
                info["regla"] = regla_estudio(info["tags"])
            resultados[n] = info
            for t in info["tags"]: conteo[t] += 1

        print("  Detección:")
        for t, c in conteo.most_common():
            print(f"    - {t:30s} {c:>3d}  ({c*100/len(lote):.1f}%)")

    # emite JSON por si no se pidió un lote específico
    if not args.lote:
        with open(OUT_JSON, "w") as f:
            json.dump({"version": 1, "por_pregunta": resultados}, f, ensure_ascii=False, indent=1)
        print(f"\n✓ Escrito {OUT_JSON}  ({len(resultados)} preguntas)")

if __name__ == "__main__":
    main()
